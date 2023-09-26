'''
Following script is written purely for learning & self help purpose.
Author : Sagar Yerva
'''

import requests
from bs4 import BeautifulSoup
import openpyxl

def WriteToExcel(rows):
    # create a new Workbook() object    
    wb = openpyxl.Workbook()
    
    # Get active sheet from the workbook  
    sheet = wb.active
    
    # Cell objects also have row, column and coordinate attributes that provide location information for the cell.
    
    # Note: The first row or column integer is 1, not 0. Cell object is created by using sheet object's cell() method.
    c1 = sheet.cell(row = 1, column = 1)
    
    # writing values to cells
    c1.value = "Name"
    
    c2 = sheet.cell(row= 1 , column = 2)
    c2.value = "Price"

    c4 = sheet.cell(row= 1 , column = 3)
    c4.value = "Per Sq Ft"

    for row in rows:
        sheet.append(row)   
    
    # even if you modify the Workbook object or its sheets and cells, the spreadsheet file will not be saved till you call the save() workbook method.
    wb.save(".\\outputs\\demo.xlsx")



base_url = "https://housing.com/in/buy/searches/"
area_ids = "P1pu1fhmtcb6ffmti" #this Housing's internal id for kharadi area in pune
full_url = base_url + area_ids

# Get Page response in HTML and parse using BeautifulSoup
page = requests.get(full_url)
soup = BeautifulSoup(page.content, "html.parser")
articles = soup.findAll('article')
excel_rows = []

# each div with an element <article> holds a property listing
for article in articles:    
    name= article.find('h2').text
    price = article.find('div', class_="css-18rodr0").text
    arr_price_per_sq_ft = article.find_all('div', class_="css-4z3njv")
    price_per_sq_ft = ""
    print(f'{name} : {price}')
    for feature in arr_price_per_sq_ft:
        price_per_sq_ft = feature.text
        print(price_per_sq_ft)
    row = [name, price, price_per_sq_ft]
    excel_rows.append(row)

WriteToExcel(excel_rows)
print('done!')